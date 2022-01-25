# -*- coding: utf-8 -*-
# @Time: 2022/1/25 9:15 AM
# @Author: wangpengpeng
# @Version: first
# @ Function:

from handle_method.handlefile_inference import *
from openpyxl import Workbook
from openpyxl import load_workbook

class HandleXlsx(HandleFile):

    def __init__(self):
        super().__init__()



    def read_method(self, file_name, sheet_name=None):
        '''

        :param file_name:需要读取的 xlsx文件
        :param sheet_name: 表名
        :return:
        '''
        data_result = []

        wb = load_workbook(file_name)
        sheets = wb.worksheets  # 获取当前所有的sheet

        # 默认获取第一张sheet
        if sheet_name is None:
            sheet1 = sheets[0]
        else:
            sheet1 = sheets[sheet_name]

        rows = sheet1.rows

        # 迭代读取所有的行
        for row in rows:
            row_val = [col.value for col in row]
            data_result.append(row_val)

        return data_result

    def write_method(self, dst_file, sheet_name='sheet',  titlename_list=[] , contents=[]):
        """

        :param dst_file: 保存的文件名称
        :param sheet_name: 表名
        :param titlename_list:标题的名字
        :param contents: 数据内容
        :return:
        """
        # 校验文件保存的文件夹是否存在
        self.check_dir(dst_file)

        try:
            wb = Workbook()
            # TODO 该处测试 index为其他值的时候 无法做到在原文件下改写
            wb.create_sheet(index=0, title=sheet_name)
            # 获取当前活跃的sheet，默认是第一个sheet
            ws = wb.active

            # 写入第一行的 标题
            for i, titlename in enumerate(titlename_list):
                ws.cell(1, i+1).value = titlename

            for content in contents:
                ws.append(content)

            wb.save(dst_file)
            return True

        except Exception as e:
            print(e, "数据保存失败 ")
            return False

